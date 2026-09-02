"""Add WhatsApp links and format phone numbers in an Excel contacts sheet."""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

INPUT = Path(r"c:\Users\Admin\Downloads\data .xlsx")
OUTPUT = Path(r"c:\Users\Admin\Downloads\data-whatsapp.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1C2438")
HEADER_FONT = Font(bold=True, color="E4C878", size=11)
LINK_FONT = Font(color="0563C1", underline="single")

KNOWN_CODES = (
    "966",
    "965",
    "971",
    "973",
    "974",
    "968",
    "962",
    "961",
    "218",
    "20",
    "212",
    "213",
    "216",
    "249",
    "967",
    "964",
    "963",
    "970",
)


def clean_raw(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value != value:
            return ""
        if value == int(value):
            value = int(value)
    s = str(value).strip()
    if "e+" in s.lower():
        try:
            s = str(int(float(s)))
        except ValueError:
            pass
    s = re.sub(r"[^\d]", "", s)
    return s


def normalize_e164(digits: str) -> str:
    if not digits:
        return ""

    for code in KNOWN_CODES:
        if digits.startswith(code) and len(digits) >= len(code) + 8:
            return digits

    if len(digits) == 10 and digits.startswith(("10", "11", "12", "15")):
        return "20" + digits

    if len(digits) == 11 and digits.startswith("0") and digits[1:3] in ("10", "11", "12", "15"):
        return "20" + digits[1:]

    if len(digits) == 9 and digits.startswith("5"):
        return "966" + digits

    if len(digits) == 8 and digits[0] in "569":
        return "965" + digits

    return digits


def format_display(e164: str) -> str:
    if not e164:
        return ""
    if e164.startswith("20") and len(e164) == 12:
        return f"+20 {e164[2:4]} {e164[4:8]} {e164[8:]}"
    if e164.startswith("966") and len(e164) == 12:
        return f"+966 {e164[3:5]} {e164[5:8]} {e164[8:]}"
    if e164.startswith("965") and len(e164) == 11:
        return f"+965 {e164[3:5]} {e164[5:8]} {e164[8:]}"
    if e164.startswith("971") and len(e164) == 12:
        return f"+971 {e164[3:5]} {e164[5:8]} {e164[8:]}"
    if e164.startswith("218") and len(e164) >= 12:
        return f"+218 {e164[3:5]} {e164[5:8]} {e164[8:]}"
    return "+" + e164


def whatsapp_url(e164: str) -> str:
    return f"https://wa.me/{e164}" if e164 else ""


def find_phone_col(headers: list) -> int | None:
    for i, h in enumerate(headers):
        if h is None:
            continue
        text = str(h).lower()
        if "phone" in text or "جوال" in text or "موبايل" in text or "هاتف" in text or "رقم" in text:
            return i
    return 1 if len(headers) > 1 else None


def process_sheet(ws) -> int:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0

    headers = list(rows[0])
    phone_idx = find_phone_col(headers)
    if phone_idx is None:
        return 0

    new_headers = (
        headers[: phone_idx + 1]
        + ["WhatsApp Link", "Open WhatsApp"]
        + headers[phone_idx + 1 :]
    )

    ws.delete_rows(1, ws.max_row)
    ws.append(new_headers)

    link_idx = phone_idx + 1
    btn_idx = phone_idx + 2

    count = 0
    for row in rows[1:]:
        if not any(row):
            continue
        row = list(row)
        while len(row) < len(headers):
            row.append(None)

        raw = clean_raw(row[phone_idx])
        e164 = normalize_e164(raw)
        display = format_display(e164) if e164 else raw
        url = whatsapp_url(e164)

        row[phone_idx] = display
        new_row = row[: phone_idx + 1] + [url, "WhatsApp" if url else ""] + row[phone_idx + 1 :]
        ws.append(new_row)
        count += 1

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in range(2, ws.max_row + 1):
        phone_cell = ws.cell(r, phone_idx + 1)
        phone_cell.number_format = "@"
        phone_cell.alignment = Alignment(horizontal="left")

        link_cell = ws.cell(r, link_idx + 1)
        link_cell.number_format = "@"
        link_cell.alignment = Alignment(horizontal="left")

        btn_cell = ws.cell(r, btn_idx + 1)
        url = link_cell.value
        if url:
            btn_cell.hyperlink = url
            btn_cell.value = "WhatsApp"
            btn_cell.font = LINK_FONT

    ws.column_dimensions[get_column_letter(1)].width = 30
    ws.column_dimensions[get_column_letter(phone_idx + 1)].width = 22
    ws.column_dimensions[get_column_letter(link_idx + 1)].width = 42
    ws.column_dimensions[get_column_letter(btn_idx + 1)].width = 14
    if len(new_headers) > btn_idx + 1:
        ws.column_dimensions[get_column_letter(btn_idx + 2)].width = 8

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return count


def main() -> None:
    wb = openpyxl.load_workbook(INPUT)
    total = 0
    for name in wb.sheetnames:
        total += process_sheet(wb[name])
    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")
    print(f"Rows: {total} | Sheets: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()
